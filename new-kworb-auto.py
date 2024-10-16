#!/usr/bin/env python
# coding: utf-8

# In[21]:

import os
import requests
import schedule
import subprocess
import time
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re

# List of URLs to scrape
urls = [
    "https://kworb.net/spotify/country/global_daily.html",
    "https://kworb.net/spotify/country/us_daily.html",
    "https://kworb.net/spotify/country/gb_daily.html",
    "https://kworb.net/spotify/country/ca_daily.html",
    "https://kworb.net/spotify/country/au_daily.html",
    "https://kworb.net/ww/index.html",
    "https://kworb.net/charts/itunes/us.html",
    "https://kworb.net/charts/itunes/uk.html",
    "https://kworb.net/charts/itunes/au.html",
    "https://kworb.net/charts/itunes/ca.html",
    "https://kworb.net/apple_songs/index.html",
    "https://kworb.net/charts/apple_s/us.html",
    "https://kworb.net/charts/apple_s/uk.html",
    "https://kworb.net/charts/apple_s/au.html",
    "https://kworb.net/charts/apple_s/ca.html",
    "https://kworb.net/youtube/index.html",
    "https://kworb.net/youtube/insights/us_daily.html",
    "https://kworb.net/youtube/insights/uk_daily.html",
    "https://kworb.net/youtube/insights/au_daily.html",
    "https://kworb.net/youtube/insights/ca_daily.html",
    "https://kworb.net/charts/shazam/ww.html",
    "https://kworb.net/charts/shazam/us.html",
    "https://kworb.net/charts/shazam/uk.html",
    "https://kworb.net/charts/shazam/au.html",
    "https://kworb.net/charts/shazam/ca.html",
    "https://kworb.net/charts/deezer/ww.html",
    "https://kworb.net/charts/deezer/us.html",
    "https://kworb.net/charts/deezer/uk.html",
    "https://kworb.net/charts/deezer/au.html",
    "https://kworb.net/charts/deezer/ca.html"
    # Example URL that should be considered as "itunes"
    # Add more URLs as needed
]

# Function to determine if a URL should be considered as "itunes"
def is_itunes_url(url):
    # Example condition: Check if the URL domain or structure indicates itunes
    if re.search(r'/charts/itunes/', url):
        return True
    # Add more conditions as needed
    return False

# Function to determine whether a URL is considered as "apple music"
def is_apple_url(url):
    if "apple" in url:
        return True
    return False

# Function to determine whether a URL is considered as "youtube"
def is_youtube_url(url):
    if "youtube" in url:
        return True
    return False

# Function to determine whether a URL is considered as "shazam"
def is_shazam_url(url):
    if "shazam" in url:
        return True
    return False

# Function to determine whether a URL is considered as "deezer"
def is_deezer_url(url):
    if "deezer" in url:
        return True
    return False

# Fetch the current date and format it
current_date = datetime.now().strftime("%Y-%m-%d")

# Create a new Excel workbook
wb = Workbook()

# Remove the default sheet named "Sheet" if it exists
default_sheet = wb.get_sheet_by_name('Sheet')
if default_sheet:
    wb.remove(default_sheet)

# Rename the default sheet (index) to "ww itunes" if it exists
if "index" in wb.sheetnames:
    ws = wb["index"]
    ws.title = "ww itunes"

# Delete existing "ww itunes" sheet if it exists
if "ww itunes" in wb.sheetnames:
    ws_to_delete = wb["ww itunes"]
    wb.remove(ws_to_delete)

# Define a yellow fill
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Dictionary to store rows with "NEW" in column B and their sources
new_entries = []

# Create "all new entries" sheet as the first sheet
all_new_entries_ws = wb.create_sheet(title="all new entries", index=0)

# Write header for "all new entries" sheet
all_new_entries_ws.append(["Rank", "Rank Change", "Artist", "Source"])

# Loop through each URL
for URL in urls:
    try:
        page = requests.get(URL)
        soup = BeautifulSoup(page.content, "html.parser")
        
        # Find all tables on the page
        tables = soup.find_all("table")
        
        if tables:
            # Assuming the table we want is the first one, modify this if needed
            table = tables[0]
            song_rankings = table.find_all("tr")
            
            # Create a new sheet for each URL with a simplified name
            sheet_name = URL.split('/')[-1].replace('.html', '')
            
            # Append "spotify" or "itunes" to sheet_name based on URL content
            if "spotify" in URL:
                sheet_name += " spotify"
            elif is_itunes_url(URL):
                sheet_name += " itunes"
            elif is_apple_url(URL):
                sheet_name += " apple music"
            elif is_youtube_url(URL):
                sheet_name += " youtube"
            elif is_shazam_url(URL):
                sheet_name += " shazam"
            elif is_deezer_url(URL):
                sheet_name += " deezer"
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Extract header row (assuming it's the first row of the table)
            header_row = table.find("tr")
            headers = [header.text.strip() for header in header_row.find_all(["th", "td"])]
            
            # Write the header to the worksheet
            ws.append(headers)
            
            # Extract and write data to the sheet
            for song in song_rankings:
                columns = song.find_all(["td", "th"])  # Look for both td and th elements
                data_row = [column.text.strip() for column in columns]
                
                # Append the data to the sheet
                ws.append(data_row)
                
                # Highlight the row yellow if "NEW" is in column B
                if "NEW" in data_row[1]:  # Check if "NEW" is in column B
                    for cell in ws[ws.max_row]:  # Highlight entire row
                        cell.fill = yellow_fill
                        
                    # Append row with source to new_entries list
                    new_entries.append([data_row[0], data_row[1], data_row[2], sheet_name])
                    
        else:
            print(f"No tables found for URL: {URL}")

    except Exception as e:
        print(f"Error processing URL {URL}: {str(e)}")

# Write rows with "NEW" in column B to "all new entries" sheet
for row in new_entries:
    all_new_entries_ws.append([row[0], row[1], row[2], row[3]])

# Define the Excel file name with the current date
excel_file_name = f"kworb_charts_{current_date}.xlsx"

# Get the path to the desktop (modify this path according to your OS)
desktop_path = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop')
excel_file_path = os.path.join(desktop_path, excel_file_name)

# Save the workbook to the specified path
wb.save(excel_file_path)

print(f"Data has been written to {excel_file_path}")


# In[ ]:

def job():
    subprocess.run(["/usr/bin/python3", "/Users/nathanpak/desktop/music_data/new-kworb-auto.py"])
    return schedule.CancelJob  # This will cancel the job after running it once

# Schedule the job every day at 10:00 AM
schedule.every().day.at("10:00").do(job)

while True:
    schedule.run_pending()
    current_time = time.strftime("%H:%M")
    if current_time == "10:00":
        job()
        break
    time.sleep(60)





